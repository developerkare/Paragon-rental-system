try:
    import openpyxl
    print("openpyxl imported successfully")
    
    excel_file = r'c:\Users\Hp\Paragon-rental-system\KCNA Prep - Quiz Materials (1).xlsx'
    wb = openpyxl.load_workbook(excel_file)
    print(f"Workbook loaded: {excel_file}")
    
    sheets = wb.sheetnames
    print(f"Sheets found: {sheets}")
    
    for sheet_name in sheets:
        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} cols")
        
        # Print first 30 rows
        for i in range(1, min(31, ws.max_row + 1)):
            cells = []
            for j in range(1, ws.max_column + 1):
                val = ws.cell(i, j).value
                if val:
                    cells.append(str(val)[:50])
            if cells:
                print(f"Row {i}: {cells}")

except Exception as e:
    print(f"Error: {type(e).__name__}: {e}")
    import traceback
    traceback.print_exc()
