import openpyxl
import json

# Load the Excel file
excel_file = r'c:\Users\Hp\Paragon-rental-system\KCNA Prep - Quiz Materials (1).xlsx'
wb = openpyxl.load_workbook(excel_file)

# Get all sheet names
sheet_names = wb.sheetnames
print("Available sheets:", sheet_names)

# Extract data from all sheets
all_quiz_data = {}

for sheet_name in sheet_names:
    ws = wb[sheet_name]
    print(f"\n--- Processing sheet: {sheet_name} ---")
    
    quiz_data = []
    
    # Get all values from the sheet
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"Row {row_idx}: {row}")
        
        # Try to identify question structure
        if row and row[0]:  # If first cell has content
            print(f"  Content: {row}")
    
    all_quiz_data[sheet_name] = quiz_data

print("\n" + "="*50)
print("Sheet structure identified")
print("="*50)

# Now let's get more detailed info
for sheet_name in sheet_names:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Print first few rows with cell references
    for row_idx in range(1, min(20, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                row_data.append(f"[{cell.coordinate}]: {cell.value}")
        if row_data:
            print(f"  Row {row_idx}: {', '.join(row_data)}")
